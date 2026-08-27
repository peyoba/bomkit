"""模板驱动 openpyxl 渲染 + 公式注入防护。见 docs/02-contracts.md #3.2、#6.3、#7。

设计说明（重要，供审阅者核对是否符合预期）：
- `base_xlsx_b64` 为空时走"内置默认模板"路径：程序化生成 title/meta/header 三行，
  复刻旧工具 create_excel 的视觉样式（旧核心 671-878 行）；随后与"上传模板"路径
  共用同一套按 `columns` 映射写数据行的逻辑。
- `base_xlsx_b64` 非空时：加载用户模板，保留 `data_start_row` 之前的全部内容与
  样式，从该行起写 fixed_rows + 数据行。
- **契约含糊点（已按保守方案实现，供任务分发者复核）**：契约 7 节"导出时 multi
  行为"写"若用户已选定（selected 唯一或 manual_code 非空）则输出单行"，但
  `selected` 字段在 analyze() 输出中总是有默认推荐值（并非"是否已处理"的标志），
  无法仅凭其非空判断用户是否已确认。本实现将"已处理"严格定义为
  `manual_code` 非空（唯一无歧义的显式信号）；否则一律展开为候选多行，与旧
  工具行为一致。如需支持"仅改 selected 视为已确认"，请在契约中补充明确的
  "resolved: bool" 标志后再调整。
"""
from __future__ import annotations

import base64
import io

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .models import cell_text
from .schema import ProfileError, validate_output_template

# ── 颜色表（契约第 7 节，预览 UI 与导出 Excel 必须一致） ──
# 注意：substring（料号匹配，低置信度）不再整行铺琥珀底色——2026-08 用户
# 决定：这类行保持默认空白即可（契约文档第 7 节的琥珀色描述已过时，待
# 任务分发者修订文档）。多候选等其余状态颜色维持不变。

FILL_MATCH = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 绿：exact/model
FILL_PARAM = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # 蓝
FILL_NONE = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 红
FILL_NON_COMPONENT = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # 灰
FILL_MISSING = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黄（单元格级）
FILL_DNP = PatternFill(start_color="E6B8AF", end_color="E6B8AF", fill_type="solid")  # 暗红
FILL_MULTI_A = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
FILL_MULTI_B = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
FILL_MULTI_PARAM_A = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FILL_MULTI_PARAM_B = PatternFill(start_color="E9EFF7", end_color="E9EFF7", fill_type="solid")

_LEVEL_FILL = {
    "exact": FILL_MATCH,
    "model": FILL_MATCH,
    "param": FILL_PARAM,
    "none": FILL_NONE,
    "non_component": FILL_NON_COMPONENT,
}

# 以 = + - @ 开头的文本单元格强制 data_type='s'，防止被解读为公式（契约 6.4 未含，
# 迁移映射第 2 节第 13 条）。
_FORMULA_TRIGGER_CHARS = ("=", "+", "-", "@")


def _write_safe_text(ws, row: int, column: int, value):
    """写入来自输入文件的文本，防公式注入（旧核心 989-1008 行原样搬运）。"""
    text = cell_text(value)
    cell = ws.cell(row=row, column=column, value=text)
    if text.startswith(_FORMULA_TRIGGER_CHARS):
        cell.data_type = "s"
    return cell


def _col_letter_to_index(letter: str) -> int:
    """'A' -> 1, 'B' -> 2, ... 'AA' -> 27."""
    idx = 0
    for ch in letter:
        idx = idx * 26 + (ord(ch.upper()) - ord("A") + 1)
    return idx


def _resolve_row(item: dict, priority_prefix: str = "01.") -> dict:
    """解析单个 analyze item 在导出时应表现为单行还是多行（候选展开）。

    返回 {"expand": bool, "rows": [ {code,name,spec,status_text,fill_key} ]}
    fill_key 用于查颜色：单行用 match.level；展开行用 'multi' 或 'multi_param'。
    """
    match = item.get("match", {})
    level = match.get("level", "none")
    manual_code = match.get("manual_code")
    candidates = match.get("candidates") or []

    if manual_code:
        # 显式手改编码：唯一无歧义信号，输出单行。尝试在候选中找同 code 的
        # name/spec 做展示，找不到则名称/规格留空（用户手填的是编码本身）。
        found = next((c for c in candidates if c.get("code") == manual_code), None)
        name = found["name"] if found else ""
        spec = found["spec"] if found else ""
        return {
            "expand": False,
            "rows": [{"code": manual_code, "name": name, "spec": spec,
                      "status_text": match.get("status_text", ""), "level": level}],
        }

    if level == "multi":
        # 去重后唯一 code 时不算真正歧义（防御性分支，正常情况下 matching.py
        # 已保证 multi 级别至少有 2 个唯一 code）。
        seen = set()
        dedup = []
        for c in candidates:
            if c["code"] not in seen:
                seen.add(c["code"])
                dedup.append(c)
        if len(dedup) <= 1:
            c = dedup[0] if dedup else {"code": "", "name": "", "spec": ""}
            return {"expand": False, "rows": [{**c, "status_text": match.get("status_text", ""), "level": level}]}

        is_param = str(match.get("status_text", "")).startswith("参数匹配")
        rows = []
        for i, c in enumerate(dedup):
            rows.append({
                "code": c["code"], "name": c["name"], "spec": c["spec"],
                "status_text": f"候选{i + 1}/{len(dedup)}",
                "level": "multi_param" if is_param else "multi",
                "_alt": i % 2,
            })
        return {"expand": True, "rows": rows}

    # 非 multi、非 manual_code：单一候选（level in exact/model/substring/param）或
    # 无候选（none/non_component/skipped）。契约 6.2 中 match 对象本身不携带
    # code/name/spec，只有 candidates + selected 索引，这里按 selected 取值。
    selected_idx = match.get("selected", 0) or 0
    if candidates and 0 <= selected_idx < len(candidates):
        chosen = candidates[selected_idx]
    else:
        chosen = {"code": "", "name": "", "spec": ""}

    # 旧核心行为对齐（05-migration-map.md 边界 case 补充）：物料名称列在
    # 未匹配到真实物料名称时（none/non_component/skipped，candidates 为空），
    # 回退显示 BOM 自身的 Secondary Category（旧核心 658 行 display_name 逻辑），
    # 而不是留空——否则测试点/安装孔/未匹配元件这一列会比旧工具少一份可读信息。
    chosen_name = chosen.get("name", "")
    if not chosen_name:
        chosen_name = cell_text(item.get("fields", {}).get("category", ""))

    return {
        "expand": False,
        "rows": [{
            "code": chosen.get("code", ""), "name": chosen_name,
            "spec": chosen.get("spec", ""), "status_text": match.get("status_text", ""),
            "level": level,
        }],
    }


def _row_fill(row_info: dict) -> PatternFill | None:
    level = row_info["level"]
    if level == "multi":
        return FILL_MULTI_A if row_info.get("_alt", 0) == 0 else FILL_MULTI_B
    if level == "multi_param":
        return FILL_MULTI_PARAM_A if row_info.get("_alt", 0) == 0 else FILL_MULTI_PARAM_B
    return _LEVEL_FILL.get(level)


def _load_or_create_workbook(profile: dict):
    base_b64 = profile.get("base_xlsx_b64")
    sheet_index = profile.get("sheet_index", 0)
    if base_b64:
        try:
            raw = base64.b64decode(base_b64)
        except Exception as exc:
            raise ProfileError("BAD_TEMPLATE", f"输出模板 base64 解码失败: {exc}") from exc
        wb = load_workbook(io.BytesIO(raw))
        ws = wb.worksheets[sheet_index]
        return wb, ws, True
    wb = Workbook()
    ws = wb.active
    ws.title = "PCBA BOM表"
    return wb, ws, False


def _write_builtin_header(ws, profile: dict, meta: dict, total_cols: int, has_material: bool):
    """程序化生成 title/meta/header 三行，复刻旧工具样式（旧核心 706-751 行）。"""
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    last_col_letter = _index_to_col_letter(total_cols)

    ws.merge_cells(f"A1:{last_col_letter}1")
    title_cell = ws["A1"]
    title_cell.value = "PCBA BOM表"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25
    for col in range(1, total_cols + 1):
        ws.cell(row=1, column=col).border = thin_border

    meta_labels = {1: "物料编码", 3: "PCBA名称", 6: "PCBA型号"}
    meta_values = {
        2: meta.get("material_code", ""),
        4: meta.get("pcba_name", ""),
        7: meta.get("pcba_model", ""),
    }
    label_font = Font(bold=True)
    for col_idx, label in meta_labels.items():
        cell = ws.cell(row=2, column=col_idx, value=label)
        cell.font = label_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col_idx, val in meta_values.items():
        _write_safe_text(ws, 2, col_idx, val).alignment = Alignment(vertical="center")
    ws.merge_cells("D2:E2")
    ws.merge_cells(f"G2:{last_col_letter}2")
    ws.row_dimensions[2].height = 20
    for col in range(1, total_cols + 1):
        ws.cell(row=2, column=col).border = thin_border

    standard_headers = ["序号", "物料编码", "物料名称", "规格型号（金蝶）",
                         "位号", "数量", "封装", "厂商", "备注"]
    aux_headers = ["JLC规格", "Description"]
    if has_material:
        aux_headers.append("匹配状态")
    headers = standard_headers + aux_headers
    aux_header_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        if col_idx > 9:
            cell.fill = aux_header_fill

    col_widths = {"A": 6, "B": 16, "C": 17, "D": 47, "E": 60, "F": 8, "G": 13,
                  "H": 15, "I": 10, "J": 30, "K": 15}
    if has_material:
        col_widths["L"] = 14
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # PCB 空板行（序号 1，旧核心 774-784 行）
    row_idx = 4
    ws.cell(row=row_idx, column=1, value=1)
    _write_safe_text(ws, row_idx, 3, meta.get("pcb_name", ""))
    _write_safe_text(ws, row_idx, 4, meta.get("pcb_model", ""))
    ws.cell(row=row_idx, column=9, value="PCB")
    for col_idx in range(1, total_cols + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")
    return thin_border


def _index_to_col_letter(n: int) -> str:
    letters = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


_BUILTIN_COLUMNS = {
    "A": "seq", "B": "code", "C": "material_name", "D": "material_spec",
    "E": "designator", "F": "qty", "G": "footprint", "H": "manufacturer",
    "I": None, "J": "mpn", "K": "description", "L": "match_status",
}


def render(final_items: list[dict], output_profile: dict, meta: dict | None = None) -> bytes:
    """契约 6.3 render(final_items, output_profile, meta) -> bytes。"""
    validate_output_template(output_profile)
    meta = meta or {}

    has_material = any(it.get("match", {}).get("level") != "skipped" for it in final_items) if final_items else False

    wb, ws, is_custom_template = _load_or_create_workbook(output_profile)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    columns = output_profile.get("columns") or _BUILTIN_COLUMNS
    total_cols = max(_col_letter_to_index(c) for c in columns)
    style = output_profile.get("style", {})
    status_colors_enabled = style.get("status_colors", True)
    missing_highlight = style.get("missing_highlight", True)
    dnp_highlight = style.get("dnp_highlight", True)

    if not is_custom_template:
        thin_border = _write_builtin_header(ws, output_profile, meta, total_cols, has_material)
        row_idx = 5  # 序号从 2 开始的数据行紧接 PCB 空板行之后
        seq_start = 2
    else:
        data_start_row = output_profile.get("data_start_row", 1)
        for cell_ref, meta_key in (output_profile.get("meta_cells") or {}).items():
            ws[cell_ref] = meta.get(meta_key, "")
        row_idx = data_start_row
        seq_start = 1
        for fixed in output_profile.get("fixed_rows") or []:
            for col_letter, template in (fixed.get("cells") or {}).items():
                value = _apply_placeholders(template, meta, seq_start)
                _write_safe_text(ws, row_idx, _col_letter_to_index(col_letter), value)
            row_idx += 1
            seq_start += 1

    designator_col_letter = _designator_col(columns)

    seq = seq_start
    for item in final_items:
        fields = item.get("fields", {})
        resolved = _resolve_row(item, output_profile.get("candidate_priority_prefix", "01."))
        rows = resolved["rows"]

        for row_info in rows:
            row_fill = _row_fill(row_info) if status_colors_enabled else None
            code_missing_cell = None
            spec_missing_cell = None

            for col_letter, field_id in columns.items():
                col_idx = _col_letter_to_index(col_letter)
                value = _resolve_field_value(field_id, fields, row_info, seq)
                if field_id in ("seq", "qty"):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                else:
                    cell = _write_safe_text(ws, row_idx, col_idx, value)
                cell.border = thin_border
                if col_letter != designator_col_letter:
                    cell.alignment = Alignment(vertical="center")

                if row_fill is not None:
                    cell.fill = row_fill

                if field_id == "code":
                    code_missing_cell = cell
                elif field_id == "material_spec":
                    spec_missing_cell = cell
                elif field_id == "description" and dnp_highlight and fields.get("dnp"):
                    cell.fill = FILL_DNP

            # 单元格级"缺失"高亮优先于行底色（契约第 7 节：编码/规格缺失单元格 -> 黄）。
            has_device = bool(cell_text(fields.get("mpn", "")))
            eligible_for_missing = (
                missing_highlight and has_material and has_device
                and row_info["level"] not in ("non_component",)
            )
            if eligible_for_missing:
                if code_missing_cell is not None and not row_info.get("code"):
                    code_missing_cell.fill = FILL_MISSING
                if spec_missing_cell is not None and not row_info.get("spec"):
                    spec_missing_cell.fill = FILL_MISSING

            if designator_col_letter:
                d_cell = ws.cell(row=row_idx, column=_col_letter_to_index(designator_col_letter))
                d_cell.alignment = Alignment(vertical="center", wrap_text=True)
                des_len = len(cell_text(fields.get("designator", "")))
                ws.row_dimensions[row_idx].height = max(15, (des_len // 60 + 1) * 15)

            row_idx += 1
        seq += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _designator_col(columns: dict) -> str | None:
    for letter, field_id in columns.items():
        if field_id == "designator":
            return letter
    return None


def _resolve_field_value(field_id, fields: dict, row_info: dict, seq: int):
    if field_id is None:
        return ""
    if field_id == "seq":
        return seq
    if field_id == "code":
        return row_info.get("code", "")
    if field_id == "material_name":
        return row_info.get("name", "")
    if field_id == "material_spec":
        return row_info.get("spec", "")
    if field_id == "match_status":
        return row_info.get("status_text", "")
    if field_id == "qty":
        return fields.get("qty", "")
    return fields.get(field_id, "")


def _apply_placeholders(template: str, meta: dict, seq: int) -> str:
    if not isinstance(template, str):
        return template
    text = template.replace("{seq}", str(seq))
    for key, val in meta.items():
        text = text.replace("{" + key + "}", cell_text(val))
    return text
