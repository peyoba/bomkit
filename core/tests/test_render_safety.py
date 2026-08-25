# -*- coding: utf-8 -*-
"""迁移自旧仓库 tests/test_output_path_and_safety.py 的公式注入部分。
输出路径防覆盖部分不适用于新架构（render() 只产出字节，不写文件——见
docs/05-migration-map.md #1 迁移映射表：该职责划归 cli/ 专用）。
"""
import io

from openpyxl import load_workbook

from bomcore.api import analyze, render
from .test_api_end_to_end import BOM_PROFILE, OUTPUT_TEMPLATE


def test_formula_like_bom_fields_are_stored_as_text():
    header = ["Name", "Designator", "Quantity", "Footprint", "Device", "Manufacturer", "Comment"]
    rows = [header, ["10kΩ", "R1", "1", "R0603", "=cmd|/c calc", "+HACKED", "-1"]]
    result = analyze(rows, None, BOM_PROFILE, None)
    xlsx_bytes = render(result["items"], OUTPUT_TEMPLATE)

    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    # 第一条数据行紧跟 PCB 占位行（第 4 行）之后，即第 5 行。
    device_cell = ws.cell(row=5, column=10)
    manufacturer_cell = ws.cell(row=5, column=8)
    comment_cell = ws.cell(row=5, column=11)

    assert device_cell.value == "=cmd|/c calc"
    assert device_cell.data_type == "s"
    assert manufacturer_cell.value == "+HACKED"
    assert manufacturer_cell.data_type == "s"
    assert comment_cell.value == "-1"
    assert comment_cell.data_type == "s"


def test_normal_text_is_unaffected():
    header = ["Name", "Designator", "Quantity", "Footprint", "Device", "Manufacturer", "Comment"]
    rows = [header, ["10kΩ", "R1", "1", "R0603", "RS-03K1002FT", "FH", "10kΩ"]]
    result = analyze(rows, None, BOM_PROFILE, None)
    xlsx_bytes = render(result["items"], OUTPUT_TEMPLATE)

    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    device_cell = ws.cell(row=5, column=10)
    assert device_cell.value == "RS-03K1002FT"
