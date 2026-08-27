# -*- coding: utf-8 -*-
"""重写自旧仓库 tests/test_end_to_end.py：rows+profiles 进，xlsx 字节出，
断言用 openpyxl 重新打开校验。语义（分组、DNP、非物料、候选展开）不变。
"""
import io

from openpyxl import load_workbook

from bomcore.api import analyze, render

BOM_PROFILE = {
    "kind": "bom_input",
    "header_row_index": 0,
    "column_map": {
        "Name": "value", "Designator": "designator", "Quantity": "qty",
        "Footprint": "footprint", "Device": "mpn", "Manufacturer": "manufacturer",
        "Comment": "description", "Secondary Category": "category", "Tolerance": "tolerance",
    },
}

MATERIAL_PROFILE = {
    "kind": "material_input",
    "header_row_index": 0,
    "column_map": {"编码": "code", "名称": "name", "规格型号": "spec", "禁用状态": "status"},
}

OUTPUT_TEMPLATE = {
    "kind": "output_template",
    "base_xlsx_b64": None,
    "sheet_index": 0,
    "data_start_row": 5,
    "columns": {
        "A": "seq", "B": "code", "C": "material_name", "D": "material_spec",
        "E": "designator", "F": "qty", "G": "footprint", "H": "manufacturer",
        "I": None, "J": "mpn", "K": "description", "L": "match_status",
    },
    "style": {"status_colors": True, "missing_highlight": True, "dnp_highlight": True},
}


def _bom_rows():
    header = ["Name", "Designator", "Quantity", "Footprint", "Device",
              "Manufacturer", "Comment", "Secondary Category", "Tolerance"]
    data = [
        ["10kΩ", "R1", "1", "R0603", "RS-03K1002FT", "FH", "10kΩ", "贴片电阻", "±1%"],
        ["10kΩ", "R2", "1", "R0603", "RS-03K1002FT", "FH", "10kΩ", "贴片电阻", "±1%"],
        ["100nF", "C1", "1", "C0603", "0603B104K500NT", "FH", "100nF", "贴片电容(MLCC)", "±10%"],
        ["10kΩ-DNP", "R3", "1", "R0603", "RS-03K1002FT", "FH", "10kΩ-DNP", "贴片电阻", "±1%"],
        ["TP1_hole", "TP1", "1", "", "Test-Point", "", "", "", ""],
    ]
    return [header] + data


def _material_rows():
    header = ["编码", "名称", "规格型号", "禁用状态"]
    data = [
        ["01.0001", "贴片电阻", "10kΩ±1%/R0603(RS-03K1002FT)/FH(风华)", "否"],
        ["01.0002", "贴片电容", "100nF±10%/50V/C0603(0603B104K500NT)/FH(风华)", "否"],
    ]
    return [header] + data


def test_full_conversion_pipeline():
    result = analyze(_bom_rows(), _material_rows(), BOM_PROFILE, MATERIAL_PROFILE)
    items = result["items"]

    levels = [it["match"]["level"] for it in items]
    # 两个非 DNP 10kΩ 分到一组、DNP 组单独、电容组、非物料 hole——共 4 组。
    assert len(items) == 4
    assert levels.count("substring") == 3  # 10kΩ 组 x2 + 电容组，均为料号匹配
    assert "non_component" in levels
    assert "none" not in levels

    xlsx_bytes = render(items, OUTPUT_TEMPLATE, meta={
        "pcba_name": "TEST-PCBA", "pcb_name": "PCB空板",
        "pcb_model": "PCB-V1", "pcba_model": "PCBA型号",
    })
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active

    assert ws.cell(row=4, column=1).value == 1
    assert ws.cell(row=4, column=3).value == "PCB空板"

    statuses = [ws.cell(row=r, column=12).value for r in range(5, ws.max_row + 1)]
    assert statuses.count("料号匹配") == 3
    assert "非物料" in statuses
    assert "未匹配" not in statuses

    # 回归断言：契约 6.2 中 match 对象本身不携带 code/name/spec（只有
    # candidates+selected），render() 必须按 selected 索引取值写入编码/名称/规格列，
    # 而不是误读一个不存在的顶层 match.code 字段（曾经的真实 bug，导致编码列全空）。
    codes = [ws.cell(row=r, column=2).value for r in range(5, ws.max_row + 1)]
    names = [ws.cell(row=r, column=3).value for r in range(5, ws.max_row + 1)]
    assert "01.0001" in codes
    assert "01.0002" in codes
    assert any(n for n in names)  # 至少一行物料名称非空

    # 回归断言（2026-08 用户决定）：料号匹配（substring）行不再整行铺琥珀底色，
    # 保持默认空白；非物料行的灰色底色维持不变。
    for r in range(5, ws.max_row + 1):
        status = ws.cell(row=r, column=12).value
        fill_type = ws.cell(row=r, column=2).fill.fill_type
        if status == "料号匹配":
            assert not fill_type, f"row {r} 料号匹配行不应有行底色"
        elif status == "非物料":
            assert fill_type, f"row {r} 非物料行应保留底色"
