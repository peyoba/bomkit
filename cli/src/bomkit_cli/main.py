#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""bomkit CLI 入口（薄壳）。参数兼容旧 jlc_bom_converter.py，复用 bomcore.api。

差异说明（相对旧 CLI）：
- 旧 CLI 直接用 pandas 读 xlsx；本 CLI 用 openpyxl 读取并转换为 rows-JSON
  （字符串化规则见 docs/02-contracts.md #2），再交给 bomcore.api。
- 默认走内置 "嘉立创EDA 专业版" + "金蝶完整物料表" + "默认 PCBA 模板" 三个 Profile，
  与旧工具默认行为等价（旧核心固定假设这两种格式）。
- 输出路径的非覆盖式自动命名（旧核心 _generate_output_filename）在此保留，
  因为这是 CLI 独有职责（Web 端浏览器下载无此问题，见 05-migration-map.md #1）。
"""
from __future__ import annotations

import argparse
import json
import sys
from importlib import resources
from pathlib import Path

from openpyxl import load_workbook

from bomcore.api import analyze, render
from bomcore.schema import ProfileError


def _load_builtin_profile(filename: str) -> dict:
    data = resources.files("bomcore").joinpath("presets", filename).read_text(encoding="utf-8")
    return json.loads(data)


def _read_xlsx_as_rows(path: str) -> list[list[str]]:
    """openpyxl 读取器：str(v) if v is not None else ""（见 05-migration-map.md #3）。"""
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(["" if v is None else str(v) for v in row])
    return rows


def _generate_output_filename(input_file: str) -> str:
    """非覆盖式默认输出路径（旧核心 107-123 行原样搬运）。"""
    input_path = Path(input_file)
    base_name = f"{input_path.stem}_converted"
    output_dir = input_path.parent

    output_path = output_dir / f"{base_name}.xlsx"
    if not output_path.exists():
        return str(output_path)

    for counter in range(1, 101):
        unique_path = output_dir / f"{base_name}_{counter}.xlsx"
        if not unique_path.exists():
            return str(unique_path)

    raise RuntimeError("无法生成输出文件名，请手动指定输出文件")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="bomkit",
        description="将 EDA 导出的 BOM 转换为公司 PCBA BOM 模板格式（bomkit CLI）。",
    )
    parser.add_argument("input_file", help="BOM Excel 文件（默认按嘉立创EDA专业版预设解析）")
    parser.add_argument("-o", "--output", dest="output_file", default=None,
                         help="输出文件路径（默认在输入文件旁自动生成，避免覆盖已有文件）")
    parser.add_argument("-m", "--material", dest="material_code_file", default=None,
                         help="物料编码表 Excel 文件（可选，默认按金蝶完整物料表预设解析）")
    parser.add_argument("--pcba-name", dest="pcba_name", default="",
                         help="PCBA 名称，填入输出表格表头第2行")
    parser.add_argument("--pcba-model", dest="pcba_model", default="",
                         help="PCBA 型号，填入输出表格表头第2行")
    parser.add_argument("--pcb-name", dest="pcb_name", default="",
                         help="PCB 空板名称，填入输出表格序号1行")
    parser.add_argument("--pcb-model", dest="pcb_model", default="",
                         help="PCB 空板型号，填入输出表格序号1行")
    return parser


def main(argv: list[str] | None = None) -> int:
    argv = sys.argv[1:] if argv is None else argv
    parser = build_parser()
    args = parser.parse_args(argv)

    output_file = args.output_file or _generate_output_filename(args.input_file)

    try:
        bom_rows = _read_xlsx_as_rows(args.input_file)
        bom_profile = _load_builtin_profile("jlc_eda_bom_input.json")

        material_rows = None
        material_profile = None
        if args.material_code_file:
            material_rows = _read_xlsx_as_rows(args.material_code_file)
            material_profile = _load_builtin_profile("kingdee_material_input.json")

        result = analyze(bom_rows, material_rows, bom_profile, material_profile)

        output_profile = _load_builtin_profile("default_output_template.json")
        meta = {
            "pcba_name": args.pcba_name, "pcba_model": args.pcba_model,
            "pcb_name": args.pcb_name, "pcb_model": args.pcb_model,
            "material_code": "",
        }
        xlsx_bytes = render(result["items"], output_profile, meta)

        Path(output_file).write_bytes(xlsx_bytes)
    except ProfileError as exc:
        print(f"错误: {exc.message}")
        return 1
    except Exception as exc:  # noqa: BLE001 -- CLI 顶层兜底，需把任何异常转为用户可读错误
        print(f"错误: {exc}")
        return 1

    print(f"转换完成！输出文件: {output_file}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
